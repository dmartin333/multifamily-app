"""
Tests for reporting exporter functions.
"""

import pytest
import openpyxl
from io import BytesIO
from src.backend.reporting.exporter import to_xlsx, to_pdf, to_pptx


class TestExporter:
    """Test cases for exporter functions."""
    
    def test_to_xlsx_returns_bytes(self):
        """Test that to_xlsx returns bytes."""
        test_data = {
            "summary": {
                "Property Name": "Test Property",
                "IRR": "12.5%",
                "NPV": "$1,000,000"
            },
            "financial_metrics": {
                "Cap Rate": "6.5%",
                "Cash on Cash": "8.2%"
            }
        }
        result = to_xlsx(test_data)
        assert isinstance(result, bytes)
        assert len(result) > 0
    
    def test_to_xlsx_with_list_data(self):
        """Test that to_xlsx handles list data correctly."""
        test_data = {
            "summary": [
                {"Year": 1, "Revenue": 1000000, "Expenses": 600000},
                {"Year": 2, "Revenue": 1050000, "Expenses": 620000}
            ]
        }
        result = to_xlsx(test_data)
        assert isinstance(result, bytes)
        assert len(result) > 0
    
    def test_to_xlsx_empty_data(self):
        """Test that to_xlsx handles empty data."""
        empty_data = {}
        result = to_xlsx(empty_data)
        assert isinstance(result, bytes)
        assert len(result) > 0
    
    def test_to_xlsx_creates_multiple_sheets(self):
        """Test that to_xlsx creates multiple sheets for different data sections."""
        test_data = {
            "summary": {
                "Property Name": "Sunset Apartments",
                "Units": "150",
                "Purchase Price": "$15,000,000"
            },
            "financial_metrics": {
                "IRR": "14.5%",
                "NPV": "$2,500,000",
                "Cap Rate": "6.5%"
            },
            "cash_flow": [
                {"Year": 1, "Revenue": 1800000, "Expenses": 1200000, "NOI": 600000},
                {"Year": 2, "Revenue": 1850000, "Expenses": 1220000, "NOI": 630000}
            ]
        }
        
        result = to_xlsx(test_data)
        
        # Load the generated Excel file to verify structure
        workbook = openpyxl.load_workbook(BytesIO(result))
        
        # Check that expected sheets exist
        assert "Summary" in workbook.sheetnames
        assert "Financial Metrics" in workbook.sheetnames
        assert "Cash Flow" in workbook.sheetnames
        
        # Check Summary sheet content
        summary_sheet = workbook["Summary"]
        assert summary_sheet['A1'].value == "Property Summary"
        assert summary_sheet['A3'].value == "Property Name"
        assert summary_sheet['B3'].value == "Sunset Apartments"
        
        # Check Financial Metrics sheet content
        metrics_sheet = workbook["Financial Metrics"]
        assert metrics_sheet['A1'].value == "Financial Metrics"
        assert metrics_sheet['A3'].value == "IRR"
        assert metrics_sheet['B3'].value == "14.5%"
        
        # Check Cash Flow sheet content
        cashflow_sheet = workbook["Cash Flow"]
        assert cashflow_sheet['A1'].value == "Cash Flow Analysis"
        assert cashflow_sheet['A3'].value == "Year"
        assert cashflow_sheet['B3'].value == "Revenue"
        assert cashflow_sheet['A4'].value == "1"
        assert cashflow_sheet['B4'].value == "1800000"
        
        workbook.close()
    
    def test_to_xlsx_with_styling(self):
        """Test that to_xlsx applies proper styling to cells."""
        test_data = {
            "summary": {
                "Property Name": "Test Property",
                "IRR": "12.5%"
            }
        }
        
        result = to_xlsx(test_data)
        workbook = openpyxl.load_workbook(BytesIO(result))
        summary_sheet = workbook["Summary"]
        
        # Check that header is styled
        header_cell = summary_sheet['A1']
        assert header_cell.font.bold is True
        assert header_cell.font.size == 14
        
        # Check that key column is styled
        key_cell = summary_sheet['A3']
        assert key_cell.font.bold is True
        
        workbook.close()
    
    def test_to_xlsx_column_width_adjustment(self):
        """Test that to_xlsx adjusts column widths appropriately."""
        test_data = {
            "summary": {
                "Very Long Property Name That Exceeds Normal Width": "Test Property",
                "IRR": "12.5%"
            }
        }
        
        result = to_xlsx(test_data)
        workbook = openpyxl.load_workbook(BytesIO(result))
        summary_sheet = workbook["Summary"]
        
        # Check that column width is adjusted (should be reasonable, not too wide)
        column_width = summary_sheet.column_dimensions['A'].width
        assert column_width > 0
        assert column_width <= 50  # Should be capped at 50
        
        workbook.close()
    
    def test_to_pdf_not_implemented(self):
        """Test that to_pdf raises NotImplementedError."""
        test_data = {
            "property_name": "Test Property",
            "irr": 0.12,
            "npv": 1000000
        }
        with pytest.raises(NotImplementedError):
            to_pdf(test_data)
    
    def test_to_pptx_not_implemented(self):
        """Test that to_pptx raises NotImplementedError."""
        test_data = {
            "property_name": "Test Property",
            "irr": 0.12,
            "npv": 1000000
        }
        with pytest.raises(NotImplementedError):
            to_pptx(test_data)
    
    def test_exporter_with_complex_data(self):
        """Test exporter functions with complex data structure."""
        complex_data = {
            "property_info": {
                "name": "Sunset Apartments",
                "address": "123 Main St",
                "units": 150
            },
            "summary": {
                "Purchase Price": "$15,000,000",
                "IRR": "14.5%",
                "NPV": "$2,500,000"
            },
            "financial_metrics": {
                "irr": 0.145,
                "npv": 2500000,
                "cap_rate": 0.065
            },
            "cash_flow": [
                {"year": 1, "revenue": 1800000, "expenses": 1200000, "noi": 600000},
                {"year": 2, "revenue": 1850000, "expenses": 1220000, "noi": 630000}
            ]
        }
        
        # Test XLSX export
        xlsx_result = to_xlsx(complex_data)
        assert isinstance(xlsx_result, bytes)
        assert len(xlsx_result) > 0
        
        # Test PDF export (should raise NotImplementedError)
        with pytest.raises(NotImplementedError):
            to_pdf(complex_data)
        
        # Test PPTX export (should raise NotImplementedError)
        with pytest.raises(NotImplementedError):
            to_pptx(complex_data) 