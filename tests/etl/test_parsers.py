"""
Tests for ETL parsers module.
"""

import pytest
import tempfile
import os
import csv
import openpyxl
from pathlib import Path
from backend.etl.parsers import parse_csv, parse_xlsx, parse_pdf


class TestParsers:
    """Test cases for ETL parser functions."""
    
    def test_parse_csv_basic(self):
        """Test basic CSV parsing with headers."""
        # Create a temporary CSV file
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp_file:
            writer = csv.writer(tmp_file)
            writer.writerow(['Property Name', 'Units', 'Purchase Price'])
            writer.writerow(['Sunset Apartments', '150', '15000000'])
            writer.writerow(['Ocean View', '200', '25000000'])
            tmp_file_path = tmp_file.name
        
        try:
            result = parse_csv(tmp_file_path)
            
            expected = [
                {'property_name': 'Sunset Apartments', 'units': '150', 'purchase_price': '15000000'},
                {'property_name': 'Ocean View', 'units': '200', 'purchase_price': '25000000'}
            ]
            
            assert result == expected
        finally:
            os.unlink(tmp_file_path)
    
    def test_parse_csv_no_headers(self):
        """Test CSV parsing without headers."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp_file:
            writer = csv.writer(tmp_file)
            writer.writerow(['Sunset Apartments', '150', '15000000'])
            writer.writerow(['Ocean View', '200', '25000000'])
            tmp_file_path = tmp_file.name
        
        try:
            result = parse_csv(tmp_file_path)
            
            expected = [
                {'column_0': 'Sunset Apartments', 'column_1': '150', 'column_2': '15000000'},
                {'column_0': 'Ocean View', 'column_1': '200', 'column_2': '25000000'}
            ]
            
            assert result == expected
        finally:
            os.unlink(tmp_file_path)
    
    def test_parse_csv_empty_file(self):
        """Test CSV parsing with empty file."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp_file:
            tmp_file_path = tmp_file.name
        
        try:
            result = parse_csv(tmp_file_path)
            assert result == []
        finally:
            os.unlink(tmp_file_path)
    
    def test_parse_csv_file_not_found(self):
        """Test CSV parsing with non-existent file."""
        with pytest.raises(FileNotFoundError):
            parse_csv("nonexistent_file.csv")
    
    def test_parse_csv_with_spaces_in_headers(self):
        """Test CSV parsing with spaces in header names."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp_file:
            writer = csv.writer(tmp_file)
            writer.writerow(['Property Name', 'Total Units', 'Purchase Price'])
            writer.writerow(['Sunset Apartments', '150', '15000000'])
            tmp_file_path = tmp_file.name
        
        try:
            result = parse_csv(tmp_file_path)
            
            expected = [
                {'property_name': 'Sunset Apartments', 'total_units': '150', 'purchase_price': '15000000'}
            ]
            
            assert result == expected
        finally:
            os.unlink(tmp_file_path)
    
    def test_parse_xlsx_basic(self):
        """Test basic XLSX parsing."""
        # Create a temporary XLSX file
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Add headers
        worksheet['A1'] = 'Property Name'
        worksheet['B1'] = 'Units'
        worksheet['C1'] = 'Purchase Price'
        
        # Add data
        worksheet['A2'] = 'Sunset Apartments'
        worksheet['B2'] = 150
        worksheet['C2'] = 15000000
        
        worksheet['A3'] = 'Ocean View'
        worksheet['B3'] = 200
        worksheet['C3'] = 25000000
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            workbook.save(tmp_file.name)
            tmp_file_path = tmp_file.name
            workbook.close()
        
        try:
            result = parse_xlsx(tmp_file_path)
            
            expected = [
                {'property_name': 'Sunset Apartments', 'units': '150', 'purchase_price': '15000000'},
                {'property_name': 'Ocean View', 'units': '200', 'purchase_price': '25000000'}
            ]
            
            assert result == expected
        finally:
            os.unlink(tmp_file_path)
    
    def test_parse_xlsx_with_empty_cells(self):
        """Test XLSX parsing with empty cells."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Add headers
        worksheet['A1'] = 'Property Name'
        worksheet['B1'] = 'Units'
        worksheet['C1'] = 'Purchase Price'
        
        # Add data with empty cells
        worksheet['A2'] = 'Sunset Apartments'
        worksheet['B2'] = 150
        # C2 is empty
        
        worksheet['A3'] = 'Ocean View'
        # B3 is empty
        worksheet['C3'] = 25000000
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            workbook.save(tmp_file.name)
            tmp_file_path = tmp_file.name
            workbook.close()
        
        try:
            result = parse_xlsx(tmp_file_path)
            
            expected = [
                {'property_name': 'Sunset Apartments', 'units': '150', 'purchase_price': None},
                {'property_name': 'Ocean View', 'units': None, 'purchase_price': '25000000'}
            ]
            
            assert result == expected
        finally:
            os.unlink(tmp_file_path)
    
    def test_parse_xlsx_file_not_found(self):
        """Test XLSX parsing with non-existent file."""
        with pytest.raises(FileNotFoundError):
            parse_xlsx("nonexistent_file.xlsx")
    
    def test_parse_xlsx_with_spaces_in_headers(self):
        """Test XLSX parsing with spaces in header names."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Add headers with spaces
        worksheet['A1'] = 'Property Name'
        worksheet['B1'] = 'Total Units'
        worksheet['C1'] = 'Purchase Price'
        
        # Add data
        worksheet['A2'] = 'Sunset Apartments'
        worksheet['B2'] = 150
        worksheet['C2'] = 15000000
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            workbook.save(tmp_file.name)
            tmp_file_path = tmp_file.name
            workbook.close()
        
        try:
            result = parse_xlsx(tmp_file_path)
            
            expected = [
                {'property_name': 'Sunset Apartments', 'total_units': '150', 'purchase_price': '15000000'}
            ]
            
            assert result == expected
        finally:
            os.unlink(tmp_file_path)
    
    def test_parse_xlsx_empty_worksheet(self):
        """Test XLSX parsing with empty worksheet."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            workbook.save(tmp_file.name)
            tmp_file_path = tmp_file.name
            workbook.close()
        
        try:
            with pytest.raises(ValueError, match="No header row found"):
                parse_xlsx(tmp_file_path)
        finally:
            os.unlink(tmp_file_path)
    
    def test_parse_pdf_not_implemented(self):
        """Test that parse_pdf raises NotImplementedError."""
        with pytest.raises(NotImplementedError) as exc_info:
            parse_pdf("test.pdf")
        assert "PDF parsing not yet implemented" in str(exc_info.value)
    
    def test_parse_csv_with_different_delimiters(self):
        """Test CSV parsing with different delimiters."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp_file:
            writer = csv.writer(tmp_file, delimiter=';')
            writer.writerow(['Property Name', 'Units', 'Purchase Price'])
            writer.writerow(['Sunset Apartments', '150', '15000000'])
            tmp_file_path = tmp_file.name
        
        try:
            result = parse_csv(tmp_file_path)
            
            expected = [
                {'property_name': 'Sunset Apartments', 'units': '150', 'purchase_price': '15000000'}
            ]
            
            assert result == expected
        finally:
            os.unlink(tmp_file_path) 