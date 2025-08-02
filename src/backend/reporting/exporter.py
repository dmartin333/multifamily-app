"""
Reporting exporter for generating various output formats.
"""

from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO


def to_xlsx(data: Dict[str, Any]) -> bytes:
    """
    Export data to Excel (XLSX) format.
    
    Args:
        data: Dictionary containing report data
        
    Returns:
        Bytes representing the Excel file
        
    Raises:
        ValueError: If data format is invalid
    """
    try:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet and create our own
        workbook.remove(workbook.active)
        
        # Create Summary sheet
        summary_sheet = workbook.create_sheet("Summary")
        
        # Write summary data if available
        if "summary" in data:
            summary = data["summary"]
            if isinstance(summary, dict):
                # Add header
                summary_sheet['A1'] = "Property Summary"
                summary_sheet['A1'].font = Font(bold=True, size=14)
                summary_sheet['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Write key-value pairs
                row = 3
                for key, value in summary.items():
                    summary_sheet[f"A{row}"] = str(key)
                    summary_sheet[f"B{row}"] = str(value)
                    # Style the key column
                    summary_sheet[f"A{row}"].font = Font(bold=True)
                    row += 1
            elif isinstance(summary, list):
                # Add header
                summary_sheet['A1'] = "Summary Data"
                summary_sheet['A1'].font = Font(bold=True, size=14)
                
                # Write list data
                for row_idx, item in enumerate(summary, 3):
                    if isinstance(item, dict):
                        for col_idx, (key, value) in enumerate(item.items(), 1):
                            if row_idx == 3:  # Header row
                                summary_sheet.cell(row=row_idx, column=col_idx, value=str(key))
                                summary_sheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)
                            else:
                                summary_sheet.cell(row=row_idx, column=col_idx, value=str(value))
                    else:
                        summary_sheet.cell(row=row_idx, column=1, value=str(item))
        
        # Create Financial Metrics sheet
        if "financial_metrics" in data:
            metrics_sheet = workbook.create_sheet("Financial Metrics")
            metrics = data["financial_metrics"]
            
            # Add header
            metrics_sheet['A1'] = "Financial Metrics"
            metrics_sheet['A1'].font = Font(bold=True, size=14)
            metrics_sheet['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            if isinstance(metrics, dict):
                row = 3
                for key, value in metrics.items():
                    metrics_sheet[f"A{row}"] = str(key)
                    metrics_sheet[f"B{row}"] = str(value)
                    # Style the key column
                    metrics_sheet[f"A{row}"].font = Font(bold=True)
                    row += 1
        
        # Create Cash Flow sheet if data exists
        if "cash_flow" in data:
            cashflow_sheet = workbook.create_sheet("Cash Flow")
            cashflow = data["cash_flow"]
            
            # Add header
            cashflow_sheet['A1'] = "Cash Flow Analysis"
            cashflow_sheet['A1'].font = Font(bold=True, size=14)
            cashflow_sheet['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            if isinstance(cashflow, list) and len(cashflow) > 0:
                # Write headers
                headers = list(cashflow[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cashflow_sheet.cell(row=3, column=col_idx, value=str(header))
                    cashflow_sheet.cell(row=3, column=col_idx).font = Font(bold=True)
                
                # Write data
                for row_idx, item in enumerate(cashflow, 4):
                    for col_idx, header in enumerate(headers, 1):
                        value = item.get(header, "")
                        cashflow_sheet.cell(row=row_idx, column=col_idx, value=str(value))
        
        # Auto-adjust column widths
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        raise ValueError(f"Failed to create Excel file: {e}")


def to_pdf(data: Dict[str, Any]) -> bytes:
    """
    Export data to PDF format.
    
    Args:
        data: Dictionary containing report data
        
    Returns:
        Bytes representing the PDF file
        
    Raises:
        ValueError: If data format is invalid
    """
    # TODO: Implement PDF export using WeasyPrint or Puppeteer
    # Example implementation:
    # from weasyprint import HTML
    # html_content = generate_html_from_data(data)
    # pdf_bytes = HTML(string=html_content).write_pdf()
    # return pdf_bytes
    
    # Placeholder implementation
    raise NotImplementedError("PDF export not yet implemented")


def to_pptx(data: Dict[str, Any]) -> bytes:
    """
    Export data to PowerPoint (PPTX) format.
    
    Args:
        data: Dictionary containing report data
        
    Returns:
        Bytes representing the PowerPoint file
        
    Raises:
        ValueError: If data format is invalid
    """
    # TODO: Implement PowerPoint export using python-pptx
    # Example implementation:
    # from pptx import Presentation
    # from pptx.util import Inches
    # prs = Presentation()
    # slide = prs.slides.add_slide(prs.slide_layouts[5])
    # title = slide.shapes.title
    # title.text = "Multifamily Underwriting Report"
    # # Add content to slides...
    # output = BytesIO()
    # prs.save(output)
    # output.seek(0)
    # return output.getvalue()
    
    # Placeholder implementation
    raise NotImplementedError("PowerPoint export not yet implemented") 