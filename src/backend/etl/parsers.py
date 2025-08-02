"""
ETL parsers for data ingestion and parsing.
"""

import csv
import openpyxl
from typing import List, Dict, Any
from pathlib import Path


def parse_csv(path: str) -> List[Dict[str, Any]]:
    """
    Parse CSV file and return list of dictionaries.
    
    Args:
        path: Path to the CSV file
        
    Returns:
        List of dictionaries representing CSV rows
        
    Raises:
        FileNotFoundError: If the CSV file doesn't exist
        ValueError: If the file is not a valid CSV
    """
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"CSV file not found: {path}")
    
    try:
        data = []
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            # Try to detect the dialect
            sample = csvfile.read(1024)
            csvfile.seek(0)
            
            # Use csv.Sniffer to detect delimiter and other CSV properties
            try:
                dialect = csv.Sniffer().sniff(sample)
                has_header = csv.Sniffer().has_header(sample)
            except csv.Error:
                # Fallback to default settings
                dialect = csv.excel
                has_header = True
            
            reader = csv.reader(csvfile, dialect)
            
            if has_header:
                # Read header row - handle empty files
                try:
                    headers = next(reader)
                except StopIteration:
                    # Empty file, return empty list
                    return []
                
                # Clean headers (remove whitespace, replace spaces with underscores)
                headers = [h.strip().replace(' ', '_').lower() for h in headers]
                
                # Read data rows
                for row in reader:
                    if row:  # Skip empty rows
                        # Create dict mapping headers to values
                        row_dict = {}
                        for i, value in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = value.strip() if value else None
                        data.append(row_dict)
            else:
                # No header, use column indices as keys
                for row in reader:
                    if row:  # Skip empty rows
                        row_dict = {f"column_{i}": value.strip() if value else None 
                                   for i, value in enumerate(row)}
                        data.append(row_dict)
        
        return data
        
    except UnicodeDecodeError:
        # Try with different encoding
        try:
            with open(file_path, 'r', newline='', encoding='latin-1') as csvfile:
                reader = csv.DictReader(csvfile)
                return [dict(row) for row in reader]
        except Exception as e:
            raise ValueError(f"Failed to parse CSV file: {e}")
    except Exception as e:
        raise ValueError(f"Failed to parse CSV file: {e}")


def parse_xlsx(path: str) -> List[Dict[str, Any]]:
    """
    Parse XLSX file and return list of dictionaries.
    
    Args:
        path: Path to the XLSX file
        
    Returns:
        List of dictionaries representing worksheet rows
        
    Raises:
        FileNotFoundError: If the XLSX file doesn't exist
        ValueError: If the file is not a valid XLSX
    """
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Get the first active sheet
        worksheet = workbook.active
        if not worksheet:
            raise ValueError("No active worksheet found in XLSX file")
        
        data = []
        headers = []
        
        # Find the first row with data (header row)
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            if any(cell.value for cell in row):
                # This is our header row
                headers = [str(cell.value).strip().replace(' ', '_').lower() 
                          if cell.value else f"column_{i+1}"
                          for i, cell in enumerate(row)]
                break
        
        if not headers:
            raise ValueError("No header row found in XLSX file")
        
        # Read data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            if any(cell.value for cell in row):  # Skip empty rows
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        # Convert cell value to appropriate type
                        value = cell.value
                        if value is not None:
                            # Convert to string for consistency
                            row_dict[headers[i]] = str(value).strip()
                        else:
                            row_dict[headers[i]] = None
                data.append(row_dict)
        
        workbook.close()
        return data
        
    except Exception as e:
        raise ValueError(f"Failed to parse XLSX file: {e}")


def parse_pdf(path: str) -> List[Dict[str, Any]]:
    """
    Parse PDF file and return list of dictionaries.
    
    Args:
        path: Path to the PDF file
        
    Returns:
        List of dictionaries representing extracted data
        
    Raises:
        NotImplementedError: Function not yet implemented
    """
    # TODO: Implement PDF parsing with OCR fallback using PyPDF2, pdfplumber, or similar
    raise NotImplementedError("PDF parsing not yet implemented") 